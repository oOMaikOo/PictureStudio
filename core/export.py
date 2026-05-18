"""
Enhanced Excel export: custom column mapping, append/overwrite mode, protocol.
"""
import csv
import json
import os
from datetime import datetime
from typing import List, Dict, Optional

from utils.logging_utils import get_logger

log = get_logger()

# Default column definitions
DEFAULT_COLUMNS = [
    {"key": "filename",        "header": "Bildname",             "enabled": True},
    {"key": "path",            "header": "Bildpfad",             "enabled": True},
    {"key": "predicted_label", "header": "Vorhergesagtes Label", "enabled": True},
    {"key": "confidence_pct",  "header": "Confidence (%)",       "enabled": True},
    {"key": "model_name",      "header": "Modell",               "enabled": True},
    {"key": "model_type",      "header": "Architektur",          "enabled": True},
    {"key": "timestamp",       "header": "Zeitstempel",          "enabled": True},
    {"key": "top1",            "header": "Top-1",                "enabled": False},
    {"key": "top2",            "header": "Top-2",                "enabled": False},
    {"key": "top3",            "header": "Top-3",                "enabled": False},
    {"key": "error",           "header": "Fehler",               "enabled": False},
]


def _get_cell_value(result: Dict, key: str, model_name: str, ts: str) -> object:
    """Resolve one column key to a cell value for an inference result row."""
    if key == "filename":
        return result.get("filename", "")
    if key == "path":
        return result.get("path", "")
    if key == "predicted_label":
        return result.get("predicted_label", "")
    if key == "confidence_pct":
        return round(result.get("confidence", 0) * 100, 2)
    if key == "model_name":
        return model_name or os.path.basename(result.get("model_path", ""))
    if key == "model_type":
        return result.get("model_type", "")
    if key == "timestamp":
        return result.get("timestamp", ts)[:19].replace("T", " ")
    if key in ("top1", "top2", "top3"):
        idx = int(key[-1]) - 1
        top_k = result.get("top_k", [])
        if idx < len(top_k):
            t = top_k[idx]
            return f"{t['label']} ({t['prob']*100:.1f}%)"
        return ""
    if key == "error":
        return result.get("error") or ""
    return result.get(key, "")


def export_results_to_excel(
    results: List[Dict],
    output_path: str,
    model_name: str = "",
    column_defs: List[Dict] = None,
    sheet_name: str = "Ergebnisse",
    append_mode: bool = False,
) -> str:
    """
    Export inference results to Excel.
    column_defs: list of {key, header, enabled} (uses DEFAULT_COLUMNS if None).
    Returns the output path.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl nicht installiert: pip install openpyxl")

    cols = [c for c in (column_defs or DEFAULT_COLUMNS) if c.get("enabled", True)]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if append_mode and os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row
        hdr_fill = PatternFill("solid", fgColor="1565C0")
        hdr_font = Font(bold=True, color="FFFFFF")
        border = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="3A3A3A"),
        )
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col["header"])
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        ws.freeze_panes = "A2"

    # Find next row
    start_row = ws.max_row + 1 if append_mode and ws.max_row > 1 else 2

    # Data rows
    fill_low = PatternFill("solid", fgColor="4A1313")   # low confidence
    fill_err = PatternFill("solid", fgColor="5C1A1A")   # error

    for ri, result in enumerate(results, start_row):
        is_error = bool(result.get("error"))
        is_low = result.get("low_confidence", False)
        for ci, col in enumerate(cols, 1):
            val = _get_cell_value(result, col["key"], model_name, ts)
            cell = ws.cell(row=ri, column=ci, value=val)
            if is_error:
                cell.fill = fill_err
            elif is_low and col["key"] == "confidence_pct":
                cell.fill = fill_low
            if col["key"] == "confidence_pct":
                cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    log.info("Excel exportiert: %s (%d Zeilen)", output_path, len(results))
    return output_path


def export_training_report(run_data: Dict, output_path: str) -> str:
    """Export a training run as a structured Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl nicht installiert.")

    wb = openpyxl.Workbook()

    def _write(ws, row, key, value, bold_key=True):
        k = ws.cell(row=row, column=1, value=key)
        if bold_key:
            k.font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))

    # Sheet: Summary
    ws = wb.active
    ws.title = "Zusammenfassung"
    metrics = run_data.get("metrics", {})
    rows = [
        ("Run-ID", run_data.get("run_id", "")),
        ("Zeitstempel", run_data.get("timestamp", "")[:19]),
        ("Architektur", run_data.get("model_type", "")),
        ("Klassen", ", ".join(run_data.get("class_names", []))),
        ("Gerät", run_data.get("device", "")),
        ("Accuracy", f"{metrics.get('accuracy', 0)*100:.2f}%"),
        ("Precision (Macro)", f"{metrics.get('macro_precision', 0)*100:.2f}%"),
        ("Recall (Macro)", f"{metrics.get('macro_recall', 0)*100:.2f}%"),
        ("F1 (Macro)", f"{metrics.get('macro_f1', 0)*100:.2f}%"),
        ("F1 (Weighted)", f"{metrics.get('weighted_f1', 0)*100:.2f}%"),
        ("Train-Samples", run_data.get("train_size", "")),
        ("Val-Samples", run_data.get("val_size", "")),
        ("Test-Samples", run_data.get("test_size", "")),
        ("Bestes Modell", run_data.get("best_model_path", "")),
        ("Early Stopped", run_data.get("early_stopped", False)),
    ]
    for i, (k, v) in enumerate(rows, 1):
        _write(ws, i, k, v)

    # Sheet: Hyperparameter
    ws_hp = wb.create_sheet("Hyperparameter")
    for i, (k, v) in enumerate(run_data.get("hyperparameters", {}).items(), 1):
        _write(ws_hp, i, k, v)

    # Sheet: Per-class metrics
    ws_cls = wb.create_sheet("Klassen-Metriken")
    headers = ["Klasse", "Precision", "Recall", "F1", "Support"]
    for ci, h in enumerate(headers, 1):
        ws_cls.cell(row=1, column=ci, value=h).font = Font(bold=True)
    for ri, (cls, v) in enumerate(metrics.get("per_class", {}).items(), 2):
        for ci, val in enumerate([
            cls, f"{v['precision']*100:.2f}%", f"{v['recall']*100:.2f}%",
            f"{v['f1']*100:.2f}%", v["support"]
        ], 1):
            ws_cls.cell(row=ri, column=ci, value=val)

    # Sheet: Training history
    ws_hist = wb.create_sheet("Trainingshistorie")
    for ci, h in enumerate(["Epoche","Train-Loss","Val-Loss","Train-Acc","Val-Acc","LR"], 1):
        ws_hist.cell(row=1, column=ci, value=h).font = Font(bold=True)
    history = run_data.get("history", {})
    n_ep = len(history.get("train_loss", []))
    for ri in range(n_ep):
        ws_hist.cell(ri+2, 1, ri+1)
        ws_hist.cell(ri+2, 2, history.get("train_loss", [])[ri] if ri < len(history.get("train_loss",[])) else "")
        ws_hist.cell(ri+2, 3, history.get("val_loss", [])[ri] if ri < len(history.get("val_loss",[])) else "")
        ws_hist.cell(ri+2, 4, history.get("train_acc", [])[ri] if ri < len(history.get("train_acc",[])) else "")
        ws_hist.cell(ri+2, 5, history.get("val_acc", [])[ri] if ri < len(history.get("val_acc",[])) else "")
        ws_hist.cell(ri+2, 6, history.get("lr", [])[ri] if ri < len(history.get("lr",[])) else "")

    # Sheet: Software versions
    ws_ver = wb.create_sheet("Softwareversionen")
    for i, (k, v) in enumerate(run_data.get("software_versions", {}).items(), 1):
        _write(ws_ver, i, k, v)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    log.info("Trainingsbericht exportiert: %s", output_path)
    return output_path


def export_results_to_csv(
    results: List[Dict],
    output_path: str,
    model_name: str = "",
    column_defs: List[Dict] = None,
) -> str:
    """Export inference results to CSV. Returns the output path."""
    cols = [c for c in (column_defs or DEFAULT_COLUMNS) if c.get("enabled", True)]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([c["header"] for c in cols])
        for result in results:
            writer.writerow([
                _get_cell_value(result, c["key"], model_name, ts) for c in cols
            ])
    log.info("CSV exportiert: %s (%d Zeilen)", output_path, len(results))
    return output_path


def export_results_to_json(
    results: List[Dict],
    output_path: str,
    model_name: str = "",
    column_defs: List[Dict] = None,
) -> str:
    """Export inference results to JSON. Returns the output path."""
    cols = [c for c in (column_defs or DEFAULT_COLUMNS) if c.get("enabled", True)]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = [
        {c["header"]: _get_cell_value(result, c["key"], model_name, ts) for c in cols}
        for result in results
    ]
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(rows, fh, indent=2, ensure_ascii=False)
    log.info("JSON exportiert: %s (%d Einträge)", output_path, len(rows))
    return output_path
