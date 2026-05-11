"""
Unit + integration tests for core.export — Excel output, column mapping, append mode.
"""
import os

import pytest


SAMPLE_RESULTS = [
    {
        "filename": "img_001.jpg",
        "path": "/fake/img_001.jpg",
        "predicted_label": "gut",
        "confidence": 0.92,
        "confidence_pct": 92.0,
        "model_name": "resnet18",
        "model_type": "resnet18",
        "timestamp": "2025-01-01T12:00:00",
        "top1": "gut (92%)",
        "top2": "neutral (5%)",
        "top3": "schlecht (3%)",
        "low_confidence": False,
        "error": "",
    },
    {
        "filename": "img_002.jpg",
        "path": "/fake/img_002.jpg",
        "predicted_label": "schlecht",
        "confidence": 0.61,
        "confidence_pct": 61.0,
        "model_name": "resnet18",
        "model_type": "resnet18",
        "timestamp": "2025-01-01T12:00:01",
        "top1": "schlecht (61%)",
        "top2": "gut (30%)",
        "top3": "neutral (9%)",
        "low_confidence": True,
        "error": "",
    },
]


def _skip_if_no_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl nicht installiert")


# ---------------------------------------------------------------------------
# export_results_to_excel
# ---------------------------------------------------------------------------

class TestExcelExport:
    def test_creates_file(self, tmp_dir):
        _skip_if_no_openpyxl()
        from core.export import export_results_to_excel
        out = os.path.join(tmp_dir, "results.xlsx")
        export_results_to_excel(SAMPLE_RESULTS, out)
        assert os.path.exists(out)

    def test_row_count(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_results_to_excel
        out = os.path.join(tmp_dir, "rows.xlsx")
        export_results_to_excel(SAMPLE_RESULTS, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # 1 header + 2 data rows
        assert ws.max_row >= 3

    def test_custom_sheet_name(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_results_to_excel
        out = os.path.join(tmp_dir, "sheet.xlsx")
        export_results_to_excel(SAMPLE_RESULTS, out, sheet_name="TestBlatt")
        wb = openpyxl.load_workbook(out)
        assert "TestBlatt" in wb.sheetnames

    def test_disabled_column_excluded(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_results_to_excel, DEFAULT_COLUMNS
        col_defs = [dict(c) for c in DEFAULT_COLUMNS]
        for c in col_defs:
            if c["key"] == "path":
                c["enabled"] = False

        out = os.path.join(tmp_dir, "cols.xlsx")
        export_results_to_excel(SAMPLE_RESULTS, out, column_defs=col_defs)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        # "Pfad" or "path" header should not appear
        assert not any(h and "pfad" in str(h).lower() for h in headers)

    def test_append_mode_adds_rows(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_results_to_excel
        out = os.path.join(tmp_dir, "append.xlsx")
        # First write
        export_results_to_excel(SAMPLE_RESULTS[:1], out, sheet_name="Data", append_mode=False)
        wb = openpyxl.load_workbook(out)
        rows_before = wb["Data"].max_row

        # Append
        export_results_to_excel(SAMPLE_RESULTS[1:], out, sheet_name="Data", append_mode=True)
        wb2 = openpyxl.load_workbook(out)
        rows_after = wb2["Data"].max_row
        assert rows_after > rows_before

    def test_custom_header_name(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_results_to_excel, DEFAULT_COLUMNS
        col_defs = [dict(c) for c in DEFAULT_COLUMNS]
        for c in col_defs:
            if c["key"] == "filename":
                c["header"] = "Bilddatei"

        out = os.path.join(tmp_dir, "header.xlsx")
        export_results_to_excel(SAMPLE_RESULTS, out, column_defs=col_defs)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        assert "Bilddatei" in headers


# ---------------------------------------------------------------------------
# export_training_report
# ---------------------------------------------------------------------------

class TestTrainingReport:
    SAMPLE_RUN = {
        "run_id": "run_test_001",
        "architecture": "resnet18",
        "class_names": ["gut", "schlecht"],
        "best_epoch": 5,
        "best_val_acc": 0.92,
        "metrics": {
            "accuracy": 0.92, "f1": 0.91,
            "class_report": {"gut": {"precision": 0.9, "recall": 0.95, "f1-score": 0.92, "support": 10},
                             "schlecht": {"precision": 0.94, "recall": 0.89, "f1-score": 0.91, "support": 10}},
            "confusion_matrix": [[9, 1], [2, 8]],
        },
        "history": {
            "train_loss": [0.9, 0.7, 0.5],
            "val_loss": [1.0, 0.8, 0.6],
            "train_acc": [0.6, 0.75, 0.85],
            "val_acc": [0.55, 0.70, 0.80],
            "lr": [0.001, 0.001, 0.0005],
        },
        "config": {"lr": 0.001, "batch_size": 16, "epochs": 10},
        "software_versions": {"python": "3.11", "torch": "2.0"},
        "started_at": "2025-01-01T10:00:00",
        "finished_at": "2025-01-01T10:30:00",
        "duration_seconds": 1800,
    }

    def test_creates_file(self, tmp_dir):
        _skip_if_no_openpyxl()
        from core.export import export_training_report
        out = os.path.join(tmp_dir, "report.xlsx")
        export_training_report(self.SAMPLE_RUN, out)
        assert os.path.exists(out)

    def test_has_multiple_sheets(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_training_report
        out = os.path.join(tmp_dir, "sheets.xlsx")
        export_training_report(self.SAMPLE_RUN, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 2

    def test_history_sheet_present(self, tmp_dir):
        _skip_if_no_openpyxl()
        import openpyxl
        from core.export import export_training_report
        out = os.path.join(tmp_dir, "hist.xlsx")
        export_training_report(self.SAMPLE_RUN, out)
        wb = openpyxl.load_workbook(out)
        names_lower = [s.lower() for s in wb.sheetnames]
        assert any("histor" in n or "verlauf" in n or "training" in n for n in names_lower)


# ---------------------------------------------------------------------------
# DEFAULT_COLUMNS schema
# ---------------------------------------------------------------------------

class TestDefaultColumns:
    def test_default_columns_is_list(self):
        from core.export import DEFAULT_COLUMNS
        assert isinstance(DEFAULT_COLUMNS, list)
        assert len(DEFAULT_COLUMNS) >= 5

    def test_each_column_has_required_fields(self):
        from core.export import DEFAULT_COLUMNS
        for col in DEFAULT_COLUMNS:
            assert "key" in col
            assert "header" in col
            assert "enabled" in col

    def test_filename_column_enabled_by_default(self):
        from core.export import DEFAULT_COLUMNS
        fn_cols = [c for c in DEFAULT_COLUMNS if c["key"] == "filename"]
        assert fn_cols and fn_cols[0]["enabled"] is True
